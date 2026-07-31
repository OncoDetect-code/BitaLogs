"""
matriz_excel.py — Genera la Matriz de Impacto (UNITEC) como archivo
Excel identico al formato institucional original:

  - Fila 1: FACULTAD DE INGENIERIA (merge A1:F1, centrado, Poppins 8 bold)
  - Fila 2: INGENIERIA EN BIOMEDICA (merge A2:F2)
  - Logo UNITEC anclado en la parte superior derecha del encabezado.
  - Fila 5: encabezados de las 7 columnas.
  - Filas 6+: los registros, con todo el texto centrado (horizontal y
    vertical) y ajuste de linea.

Expone:
    matriz_xlsx_bytes(registros, titulo) -> bytes
"""

import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage

_AZUL = "1F4E78"
_GRIS_ENC = "D9E1F2"
_FUENTE = "Poppins"   # si el equipo no la tiene instalada, Excel usa un fallback

_LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo_unitec.png")

# Anchos de columna tomados del Excel original.
_ANCHOS = {"A": 11.2, "B": 26.4, "C": 39.9, "D": 34.2,
           "E": 15.8, "F": 34.0, "G": 26.9}

_ENCABEZADOS = ["Semana", "Equipo o proceso", "Problema identificado",
                "Solucion sugerida", "Problema ya resuelto?",
                "Impacto esperado o beneficio real", "Observaciones"]

_CLAVES = ["_semana", "equipo", "problema", "solucion",
           "resuelto", "impacto", "observaciones"]


def _val(rep, key):
    v = rep.get(key)
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan", "") else s


def matriz_xlsx_bytes(registros, titulo="Matriz de Impacto"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz de impacto"

    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    f_titulo = Font(name=_FUENTE, size=8, bold=True)
    f_enc = Font(name=_FUENTE, size=8, bold=True)
    f_dato = Font(name=_FUENTE, size=8, bold=False)
    borde = Border(*(4 * (Side(style="thin", color="444444"),)))
    fill_enc = PatternFill("solid", fgColor=_GRIS_ENC)

    # Anchos de columna
    for col, w in _ANCHOS.items():
        ws.column_dimensions[col].width = w

    # --- Encabezado institucional (merges A1:F1 y A2:F2) ---
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws["A1"] = "FACULTAD DE INGENIERIA"
    ws["A2"] = "INGENIERIA EN BIOMEDICA"
    for coord in ("A1", "A2"):
        ws[coord].font = f_titulo
        ws[coord].alignment = centro
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 15

    # Subtitulo (fila 3, merge A3:F3)
    ws.merge_cells("A3:F3")
    ws["A3"] = titulo
    ws["A3"].font = Font(name=_FUENTE, size=8, bold=False, color=_AZUL)
    ws["A3"].alignment = centro

    # --- Logo en la parte superior derecha (columna G, filas 1-3) ---
    if os.path.exists(_LOGO_PATH):
        try:
            logo = XLImage(_LOGO_PATH)
            # Escalar a un tamano prolijo dentro del area del encabezado.
            logo.width = 150
            logo.height = 84
            ws.add_image(logo, "G1")
            ws.row_dimensions[1].height = 22
        except Exception:
            pass

    # --- Fila de encabezados de columna (fila 5) ---
    fila_enc = 5
    for i, texto in enumerate(_ENCABEZADOS):
        c = ws.cell(row=fila_enc, column=i + 1, value=texto)
        c.font = f_enc
        c.alignment = centro
        c.fill = fill_enc
        c.border = borde
    ws.row_dimensions[fila_enc].height = 27

    # --- Registros (fila 6 en adelante) ---
    fila = fila_enc + 1
    if not registros:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
        c = ws.cell(row=fila, column=1,
                    value="Sin registros para el filtro seleccionado.")
        c.font = f_dato
        c.alignment = centro
        c.border = borde
    else:
        for rep in registros:
            sem = _val(rep, "semana")
            dia = _val(rep, "dia")
            rep = dict(rep)
            rep["_semana"] = f"Semana {sem}" + (f" - Dia {dia}" if dia else "")
            for i, clave in enumerate(_CLAVES):
                c = ws.cell(row=fila, column=i + 1, value=_val(rep, clave))
                c.font = f_dato
                c.alignment = centro   # todo centrado, como se pidio
                c.border = borde
            ws.row_dimensions[fila].height = 60
            fila += 1

    # Configuracion de impresion: apaisado y ajustado al ancho de una
    # pagina, para que la tabla no se parta al imprimir o exportar a PDF.
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = f"A1:G{(fila - 1) if registros else fila}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
