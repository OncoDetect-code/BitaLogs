"""
BitaLogs - Dashboard de rendimiento de práctica profesional.
Ingeniería Biomédica · UNITEC · Autor: Luis

Basado en la estructura de ServiDox, adaptado para medir el rendimiento
del practicante: equipos atendidos, tipo de mantenimiento, área, horas
acumuladas, duración y promedio de atención, con carga de la bitácora
institucional (Matriz de impacto) y comentarios semanales de evaluadores.

CÓMO CORRERLO (terminal de Windows, una sola vez):
    pip install -r requirements.txt
    python -m streamlit run bitalogs.py

Se abre en el navegador (http://localhost:8501).

BASE DE DATOS (PostgreSQL / Supabase):
La conexión se lee de st.secrets["DB_URL"]. En local, crea el archivo
.streamlit/secrets.toml con:

    DB_URL = "postgresql+psycopg://postgres.xxxx:TU_PASSWORD@aws-0-...:6543/postgres"

En Streamlit Cloud, ese mismo valor se pega en Settings -> Secrets.
Las tablas se crean solas la primera vez (init_db).
"""

from datetime import date, datetime, time, timezone, timedelta
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from sqlalchemy import create_engine, text

import calendario as cal
from importar_bitacora import leer_matriz
from formato_bitacora import bitacora_html
from matriz_excel import matriz_xlsx_bytes
import ui_dashboard as ui
from reporte_pdf import construir_pdf, construir_pdf_multi
from splash import mostrar_splash
import ai_extract

# Honduras usa UTC-6 todo el año (sin horario de verano). En Streamlit
# Cloud el servidor corre en UTC, así que date.today() daría el día
# equivocado; hoy_hn() devuelve la fecha real de Honduras.
_TZ_HN = timezone(timedelta(hours=-6))


def hoy_hn() -> date:
    return datetime.now(_TZ_HN).date()

# ------------------------------------------------------------- Catálogos
AREAS = ["Hospitalización A", "Hospitalización B", "UCI A", "UCI B",
         "Sala Cuna", "UCIN", "Emergencia", "CEYE", "Laboratorio",
         "Quirófano 1", "Quirófano 2", "Quirófano 3", "Quirófano 4",
         "Maternidad", "Diagnóstico por imágenes",
         "Área de mantenimiento", "HDV La Lima", "Otra"]
TIPOS = ["Preventivo", "Correctivo", "Revisión y Diagnóstico",
         "Instalación", "Capacitación", "Otro"]
RESUELTO = ["Sí", "Parcial", "No"]
TIPOS_ACTIVIDAD_EXTRA = ["Visita técnica", "Infografía / material educativo",
                        "Protocolo", "Revisión de equipos (sin mantenimiento)",
                        "Reunión / capacitación recibida", "Otro"]

# Nombres bonitos de columnas (mostrar y exportar)
COLS = {
    "id": "ID", "fecha": "Fecha", "semana": "Semana", "dia": "Día",
    "hora_inicio": "Inicio", "hora_fin": "Fin", "duracion_min": "Duración (min)",
    "area": "Área", "equipo": "Equipo", "marca": "Marca", "modelo": "Modelo",
    "serie": "Serie No.", "tipo": "Tipo de mantenimiento",
    "problema": "Problema identificado", "solucion": "Solución sugerida",
    "resuelto": "¿Resuelto?", "impacto": "Impacto / beneficio",
    "observaciones": "Observaciones",
}
COLS_EDIT = ["fecha", "hora_inicio", "hora_fin", "area", "equipo", "marca",
             "modelo", "serie", "tipo", "problema", "solucion", "resuelto",
             "impacto", "observaciones"]


# ------------------------------------------------------------- DB layer
@st.cache_resource
def get_engine():
    """
    Motor SQLAlchemy hacia PostgreSQL (Supabase). La URL vive en
    st.secrets["DB_URL"]. pool_pre_ping evita conexiones muertas del
    pooler; prepare_threshold=None desactiva los prepared statements
    que chocan con el pooler de Supabase (puerto 6543).
    """
    url = st.secrets["DB_URL"]
    return create_engine(url, pool_pre_ping=True,
                         connect_args={"prepare_threshold": None})


def init_db():
    with get_engine().begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS atenciones (
                id SERIAL PRIMARY KEY,
                fecha TEXT NOT NULL,
                semana INTEGER, dia INTEGER,
                hora_inicio TEXT, hora_fin TEXT, duracion_min REAL,
                area TEXT, equipo TEXT, marca TEXT, modelo TEXT, serie TEXT,
                tipo TEXT, problema TEXT, solucion TEXT, resuelto TEXT,
                impacto TEXT, observaciones TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS comentarios (
                id SERIAL PRIMARY KEY,
                semana INTEGER NOT NULL,
                evaluador TEXT NOT NULL,
                comentario TEXT NOT NULL,
                fecha_registro TEXT NOT NULL
            )
        """))
        # Actividades que NO son mantenimiento de un equipo puntual
        # (visitas, infografías, protocolos, reuniones...). Se registran
        # aparte porque no tienen "equipo/marca/serie", solo tiempo
        # invertido. Tabla nueva -> no afecta nada de 'atenciones'.
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS actividades_extra (
                id SERIAL PRIMARY KEY,
                fecha TEXT NOT NULL,
                semana INTEGER, dia INTEGER,
                tipo_actividad TEXT NOT NULL,
                descripcion TEXT,
                horas REAL NOT NULL
            )
        """))
        # Columnas para hasta 4 imágenes (base64). ADD COLUMN IF NOT EXISTS
        # es seguro: no toca datos existentes ni falla si ya están.
        for i in (1, 2, 3, 4):
            conn.execute(text(
                f"ALTER TABLE atenciones ADD COLUMN IF NOT EXISTS img{i} TEXT"))


def _dur_min(hi: str, hf: str):
    """Minutos entre hora_inicio y hora_fin (HH:MM). None si falta o inválido."""
    try:
        a = datetime.strptime(hi, "%H:%M")
        b = datetime.strptime(hf, "%H:%M")
        d = (b - a).total_seconds() / 60.0
        return round(d, 1) if d >= 0 else None
    except (ValueError, TypeError):
        return None


def _img_a_base64(archivo, max_lado: int = 1000, calidad: int = 72):
    """
    Convierte una imagen subida (JPEG/PNG) a un data-URI base64 comprimido,
    listo para guardar en la base y mostrar en HTML. Redimensiona el lado
    mayor a max_lado px para no inflar la base de datos. Devuelve None si
    el archivo no es válido.
    """
    import base64
    import io
    try:
        from PIL import Image
        im = Image.open(archivo)
        # Convertir a RGB (por si viene con transparencia o modo raro)
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        # Redimensionar manteniendo proporción
        w, h = im.size
        if max(w, h) > max_lado:
            escala = max_lado / max(w, h)
            im = im.resize((int(w * escala), int(h * escala)),
                           Image.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, format="JPEG", quality=calidad, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/jpeg;base64,{b64}"
    except Exception:
        return None


def insert_atencion(d: dict):
    d = dict(d)
    d["duracion_min"] = _dur_min(d.get("hora_inicio"), d.get("hora_fin"))
    with get_engine().begin() as conn:
        conn.execute(text("""
            INSERT INTO atenciones
            (fecha, semana, dia, hora_inicio, hora_fin, duracion_min, area,
             equipo, marca, modelo, serie, tipo, problema, solucion, resuelto,
             impacto, observaciones, img1, img2, img3, img4)
            VALUES (:fecha, :semana, :dia, :hora_inicio, :hora_fin,
                    :duracion_min, :area, :equipo, :marca, :modelo, :serie,
                    :tipo, :problema, :solucion, :resuelto, :impacto,
                    :observaciones, :img1, :img2, :img3, :img4)
        """), {
            "fecha": d["fecha"], "semana": d["semana"], "dia": d["dia"],
            "hora_inicio": d["hora_inicio"], "hora_fin": d["hora_fin"],
            "duracion_min": d["duracion_min"], "area": d["area"],
            "equipo": d["equipo"], "marca": d["marca"], "modelo": d["modelo"],
            "serie": d["serie"], "tipo": d["tipo"], "problema": d["problema"],
            "solucion": d["solucion"], "resuelto": d["resuelto"],
            "impacto": d["impacto"], "observaciones": d["observaciones"],
            "img1": d.get("img1"), "img2": d.get("img2"),
            "img3": d.get("img3"), "img4": d.get("img4"),
        })


def insert_muchas(regs: list[dict]):
    for r in regs:
        insert_atencion(r)


def load_atenciones() -> pd.DataFrame:
    with get_engine().connect() as conn:
        return pd.read_sql_query(
            text("SELECT * FROM atenciones ORDER BY fecha DESC, id DESC"), conn)


def insert_actividad_extra(d: dict):
    """Guarda una actividad que no es mantenimiento de equipo (solo tiempo)."""
    with get_engine().begin() as conn:
        conn.execute(text("""
            INSERT INTO actividades_extra
            (fecha, semana, dia, tipo_actividad, descripcion, horas)
            VALUES (:fecha, :semana, :dia, :tipo_actividad, :descripcion, :horas)
        """), d)


def load_actividades_extra() -> pd.DataFrame:
    with get_engine().connect() as conn:
        return pd.read_sql_query(
            text("SELECT * FROM actividades_extra ORDER BY fecha DESC, id DESC"),
            conn)


def delete_actividad_extra(id_: int):
    with get_engine().begin() as conn:
        conn.execute(text("DELETE FROM actividades_extra WHERE id = :id"),
                    {"id": id_})


def normalizar_tipo_actividad(texto: str) -> str:
    """
    Cuando el usuario escribe libremente el nombre de una actividad
    (opción 'Otro'), busca si ya existe una actividad igual -sin
    importar mayúsculas/minúsculas ni espacios extra- entre las
    categorías fijas y las ya guardadas en la base, y devuelve esa
    misma forma para que el gráfico las sume juntas en un solo
    registro en vez de crear una categoría duplicada.
    """
    texto = (texto or "").strip()
    if not texto:
        return "Otro"
    existentes = list(TIPOS_ACTIVIDAD_EXTRA)
    try:
        existentes += load_actividades_extra()["tipo_actividad"].dropna().tolist()
    except Exception:
        pass
    clave = texto.lower()
    for e in existentes:
        if str(e).strip().lower() == clave:
            return str(e).strip()
    return texto


def update_actividades_extra_from_df(df_edit: pd.DataFrame):
    """
    Aplica los cambios del editor de 'Otra actividad' (edición y
    borrado de filas), igual que se hace con las atenciones: por id,
    sin recrear toda la tabla.
    """
    with get_engine().begin() as conn:
        prev = {row[0] for row in conn.execute(
            text("SELECT id FROM actividades_extra")).fetchall()}
        vistos = set()
        for _, r in df_edit.iterrows():
            rid = r.get("id")
            tiene_id = pd.notna(rid) and str(rid).strip() not in ("", "None")
            horas = pd.to_numeric(r.get("horas"), errors="coerce")
            horas = float(horas) if pd.notna(horas) else 0.0
            sem = pd.to_numeric(r.get("semana"), errors="coerce")
            dia = pd.to_numeric(r.get("dia"), errors="coerce")
            vals = {
                "fecha": str(r.get("fecha", "")).strip()[:10],
                "semana": int(sem) if pd.notna(sem) else None,
                "dia": int(dia) if pd.notna(dia) else None,
                "tipo_actividad": normalizar_tipo_actividad(r.get("tipo_actividad")),
                "descripcion": r.get("descripcion") or "",
                "horas": horas,
            }
            if tiene_id:
                rid = int(rid)
                vistos.add(rid)
                vals["id"] = rid
                conn.execute(text("""
                    UPDATE actividades_extra
                    SET fecha=:fecha, semana=:semana, dia=:dia,
                        tipo_actividad=:tipo_actividad,
                        descripcion=:descripcion, horas=:horas
                    WHERE id=:id
                """), vals)
            else:
                conn.execute(text("""
                    INSERT INTO actividades_extra
                    (fecha, semana, dia, tipo_actividad, descripcion, horas)
                    VALUES (:fecha, :semana, :dia, :tipo_actividad,
                            :descripcion, :horas)
                """), vals)
        for rid in prev - vistos:
            conn.execute(text("DELETE FROM actividades_extra WHERE id=:id"),
                        {"id": rid})


def _orden_secuencial(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega una columna 'n' con la numeración amigable: 1, 2, 3... en
    orden cronológico de la práctica (Semana → Día → hora de inicio →
    id como desempate), sin importar el id crudo de la base. El id real
    se conserva intacto para las operaciones internas.
    """
    if df.empty:
        df = df.copy()
        df["n"] = []
        return df
    df = df.copy()
    df["_sem_o"] = pd.to_numeric(df.get("semana"), errors="coerce")
    df["_dia_o"] = pd.to_numeric(df.get("dia"), errors="coerce")
    df["_hi_o"] = df.get("hora_inicio", "").astype(str)
    orden = (df.sort_values(
                ["_sem_o", "_dia_o", "_hi_o", "id"], na_position="last")
             .reset_index())
    orden["n"] = range(1, len(orden) + 1)
    mapa = dict(zip(orden["id"], orden["n"]))
    df["n"] = df["id"].map(mapa)
    return df.drop(columns=["_sem_o", "_dia_o", "_hi_o"])


def update_atenciones_from_df(df_edit: pd.DataFrame):
    """
    Aplica los cambios del editor SIN borrar toda la tabla, para preservar
    los id y las imágenes. Para cada fila:
      - si trae id existente -> UPDATE de sus campos (no toca las imágenes)
      - si no trae id (fila nueva) -> INSERT
    Las filas que estaban y ya no aparecen se eliminan.
    Respeta semana/día editados a mano; si faltan, los deduce de la fecha.
    """
    campos = ["fecha", "semana", "dia", "hora_inicio", "hora_fin",
              "duracion_min", "area", "equipo", "marca", "modelo", "serie",
              "tipo", "problema", "solucion", "resuelto", "impacto",
              "observaciones"]

    def _norm_row(r):
        fecha = str(r.get("fecha", "")).strip()[:10]
        sem = r.get("semana"); dia = r.get("dia")
        sem = int(sem) if str(sem).strip() not in ("", "None", "nan") else None
        dia = int(dia) if str(dia).strip() not in ("", "None", "nan") else None
        if (sem is None or dia is None) and fecha:
            try:
                fd = datetime.strptime(fecha, "%Y-%m-%d").date()
                if sem is None:
                    sem = cal.semana_de_fecha(fd)
                if dia is None:
                    dia = cal.dia_de_semana_num(fd)
            except ValueError:
                pass
        hi = str(r.get("hora_inicio", "") or "")
        hf = str(r.get("hora_fin", "") or "")
        return {
            "fecha": fecha, "semana": sem, "dia": dia,
            "hora_inicio": hi, "hora_fin": hf, "duracion_min": _dur_min(hi, hf),
            "area": r.get("area"), "equipo": r.get("equipo"),
            "marca": r.get("marca"), "modelo": r.get("modelo"),
            "serie": r.get("serie"), "tipo": r.get("tipo"),
            "problema": r.get("problema"), "solucion": r.get("solucion"),
            "resuelto": r.get("resuelto"), "impacto": r.get("impacto"),
            "observaciones": r.get("observaciones"),
        }

    with get_engine().begin() as conn:
        # ids que existían antes
        prev = {row[0] for row in conn.execute(
            text("SELECT id FROM atenciones")).fetchall()}
        vistos = set()

        for _, r in df_edit.iterrows():
            fecha = str(r.get("fecha", "")).strip()
            if not fecha:
                continue
            vals = _norm_row(r)
            rid = r.get("id")
            tiene_id = str(rid).strip() not in ("", "None", "nan")
            if tiene_id:
                rid = int(float(rid))
                vistos.add(rid)
                set_clause = ", ".join(f"{c} = :{c}" for c in campos)
                conn.execute(
                    text(f"UPDATE atenciones SET {set_clause} WHERE id = :id"),
                    {**vals, "id": rid})
            else:
                # fila nueva: insertar (sin imágenes)
                cols = ", ".join(campos)
                ph = ", ".join(f":{c}" for c in campos)
                conn.execute(
                    text(f"INSERT INTO atenciones ({cols}) VALUES ({ph})"), vals)

        # eliminar las filas que el usuario quitó del editor
        borrar = prev - vistos
        for rid in borrar:
            conn.execute(text("DELETE FROM atenciones WHERE id = :id"),
                         {"id": rid})


def delete_atencion(id_: int):
    with get_engine().begin() as conn:
        conn.execute(text("DELETE FROM atenciones WHERE id = :id"), {"id": id_})


def update_imagenes(id_: int, imgs: list):
    """
    Reemplaza las 4 imágenes de un registro existente por su id.
    imgs es una lista de hasta 4 data-URIs (o None). Los None dejan
    ese slot vacío. Útil para agregar fotos a registros antiguos.
    """
    imgs = list(imgs[:4]) + [None] * (4 - len(imgs))
    with get_engine().begin() as conn:
        conn.execute(text("""
            UPDATE atenciones
            SET img1 = :img1, img2 = :img2, img3 = :img3, img4 = :img4
            WHERE id = :id
        """), {"id": id_, "img1": imgs[0], "img2": imgs[1],
               "img3": imgs[2], "img4": imgs[3]})


def reordenar_imagenes(id_: int, nuevo_orden: list):
    """
    Reordena las imágenes de un registro según 'nuevo_orden', una lista
    de data-URIs ya en el orden deseado (los que van primero se muestran
    primero). Rellena con None hasta 4 slots. No comprime ni recodifica:
    solo reasigna qué foto ocupa img1..img4.
    """
    orden = [x for x in nuevo_orden if x and str(x).strip()][:4]
    orden = orden + [None] * (4 - len(orden))
    with get_engine().begin() as conn:
        conn.execute(text("""
            UPDATE atenciones
            SET img1 = :img1, img2 = :img2, img3 = :img3, img4 = :img4
            WHERE id = :id
        """), {"id": id_, "img1": orden[0], "img2": orden[1],
               "img3": orden[2], "img4": orden[3]})


def imagenes_de(id_: int) -> list:
    """Devuelve las imágenes actuales (data-URIs no vacíos) de un registro."""
    with get_engine().connect() as conn:
        row = conn.execute(text(
            "SELECT img1, img2, img3, img4 FROM atenciones WHERE id = :id"),
            {"id": id_}).fetchone()
    if not row:
        return []
    return [x for x in row if x and str(x).strip()]


# ---- Comentarios de evaluadores
def insert_comentario(semana: int, evaluador: str, comentario: str):
    with get_engine().begin() as conn:
        conn.execute(text("""
            INSERT INTO comentarios (semana, evaluador, comentario, fecha_registro)
            VALUES (:semana, :evaluador, :comentario, :fecha_registro)
        """), {
            "semana": semana, "evaluador": evaluador.strip(),
            "comentario": comentario.strip(),
            "fecha_registro": datetime.now().isoformat(timespec="minutes"),
        })


def load_comentarios() -> pd.DataFrame:
    with get_engine().connect() as conn:
        return pd.read_sql_query(
            text("SELECT * FROM comentarios ORDER BY semana, fecha_registro"),
            conn)


def delete_comentario(id_: int):
    with get_engine().begin() as conn:
        conn.execute(text("DELETE FROM comentarios WHERE id = :id"), {"id": id_})


# ------------------------------------------------------------- Excel export
def build_excel(df: pd.DataFrame, dcom: pd.DataFrame, path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, data: pd.DataFrame, table_name: str):
        cols = list(data.columns)
        if not cols:
            return
        for j, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=j, value=col)
            c.fill = header_fill; c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border
        for i, (_, row) in enumerate(data.iterrows(), start=2):
            for j, col in enumerate(cols, 1):
                c = ws.cell(row=i, column=j, value=row[col])
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if i % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="F2F6FB")
        nrows = max(len(data) + 1, 2)
        ref = f"A1:{get_column_letter(len(cols))}{nrows}"
        if len(data) > 0:
            tbl = Table(displayName=table_name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tbl)
        for j, col in enumerate(cols, 1):
            maxlen = max([len(str(col))] +
                         [len(str(v)) for v in data[col].astype(str).tolist()[:200]]
                         or [0])
            ws.column_dimensions[get_column_letter(j)].width = min(
                max(maxlen + 2, 10), 50)
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "Atenciones"
    dshow = df.rename(columns=COLS)
    write_sheet(ws1, dshow, "TablaAtenciones")

    if not dcom.empty:
        ws2 = wb.create_sheet("Comentarios")
        cshow = dcom.rename(columns={
            "semana": "Semana", "evaluador": "Evaluador",
            "comentario": "Comentario", "fecha_registro": "Fecha"}).drop(
            columns=["id"], errors="ignore")
        write_sheet(ws2, cshow, "TablaComentarios")

    wb.save(path)


# =============================================================== UI
_ICONO = Path(__file__).parent / "bitalogs_icon.png"
st.set_page_config(page_title="BitaLogs · Práctica Profesional",
                   page_icon=str(_ICONO) if _ICONO.exists() else "📘",
                   layout="wide")

ui.inyectar_estilos()   # estilo visual del dashboard rediseñado

# CSS responsive: en desktop mantiene el ancho amplio; en celular fuerza
# que el contenido ocupe todo el ancho y quede centrado (sin el hueco a
# la derecha). Se usan selectores fuertes + !important porque Streamlit
# aplica sus propios anchos que de otro modo ganan.
st.markdown("""
<style>
    /* Desktop: contenido centrado con ancho máximo cómodo */
    .stMainBlockContainer, .block-container,
    section.main > div.block-container {
        max-width: 1400px !important;
        margin-left: auto !important;
        margin-right: auto !important;
        padding-left: 3rem !important;
        padding-right: 3rem !important;
    }

    /* Celular: ocupar TODO el ancho, sin hueco a la derecha */
    @media (max-width: 768px) {
        .stMainBlockContainer, .block-container,
        section.main > div.block-container {
            max-width: 100% !important;
            width: 100% !important;
            padding-left: 1rem !important;
            padding-right: 1rem !important;
            padding-top: 1rem !important;
            margin-left: 0 !important;
            margin-right: 0 !important;
        }
        /* El área principal completa al 100% */
        section.main, .stMain, [data-testid="stMain"] {
            width: 100% !important;
        }
        [data-testid="stAppViewContainer"] {
            width: 100% !important;
        }
        /* KPIs y título legibles en vertical */
        [data-testid="stMetricValue"] { font-size: 1.4rem !important; }
        h1 { font-size: 1.6rem !important; }
        /* Tabs con scroll horizontal si no caben */
        [data-testid="stTabs"] [data-baseweb="tab-list"] {
            overflow-x: auto;
            flex-wrap: nowrap;
        }
    }
</style>
""", unsafe_allow_html=True)

# --- Ícono para "Agregar a pantalla de inicio" (Android/iPhone) ---
# Lee bitalogs_icon.png, lo incrusta como base64 y declara las etiquetas
# que Chrome/Safari usan para el ícono del acceso directo, más un manifest
# mínimo para que Android lo trate como app (nombre + logo BitaLogs).
def _inyectar_icono_movil():
    import base64
    import json
    import streamlit.components.v1 as _components
    if not _ICONO.exists():
        return
    b64 = base64.b64encode(_ICONO.read_bytes()).decode("ascii")
    data_uri = f"data:image/png;base64,{b64}"
    manifest = {
        "name": "BitaLogs",
        "short_name": "BitaLogs",
        "icons": [{"src": data_uri, "sizes": "192x192", "type": "image/png"},
                  {"src": data_uri, "sizes": "512x512", "type": "image/png"}],
        "display": "standalone",
        "background_color": "#FFFFFF",
        "theme_color": "#1F4E78",
    }
    manifest_uri = ("data:application/manifest+json;base64," +
                    base64.b64encode(json.dumps(manifest).encode()).decode())
    # Se inyecta en el <head> del documento padre (la app real, no el iframe).
    _components.html(f"""
    <script>
    (function() {{
      var doc = window.parent.document;
      function add(tag, attrs) {{
        var el = doc.createElement(tag);
        for (var k in attrs) el.setAttribute(k, attrs[k]);
        doc.head.appendChild(el);
      }}
      if (!doc.getElementById('bl-touch-icon')) {{
        var l1 = doc.createElement('link');
        l1.id = 'bl-touch-icon';
        l1.rel = 'apple-touch-icon';
        l1.href = '{data_uri}';
        doc.head.appendChild(l1);
        add('link', {{rel: 'icon', type: 'image/png', href: '{data_uri}'}});
        add('link', {{rel: 'manifest', href: '{manifest_uri}'}});
        add('meta', {{name: 'apple-mobile-web-app-title', content: 'BitaLogs'}});
        add('meta', {{name: 'apple-mobile-web-app-capable', content: 'yes'}});
      }}
    }})();
    </script>
    """, height=0, width=0)


_inyectar_icono_movil()

mostrar_splash(st)   # pantalla de carga (~3.5 s) al abrir o recargar
init_db()

_logo_col, _tit_col = st.columns([1, 9])
with _logo_col:
    if _ICONO.exists():
        st.image(str(_ICONO), width=64)
with _tit_col:
    st.title("BitaLogs")
    st.caption("Dashboard de Rendimiento · Práctica Profesional Luis Velásquez 21941285 · "
               "Ingeniería Biomédica UNITEC Q3 2026")

# Botón para volver a ver la animación de carga cuando se quiera
with st.sidebar:
    if st.button("▶️ Ver animación de carga"):
        mostrar_splash(st, forzar=True)

tab_dash, tab_input, tab_carga, tab_datos, tab_bitacora, tab_coment = st.tabs(
    ["📊 Dashboard", "➕ Nuevo registro", "📥 Cargar bitácora",
     "🗂️ Datos", "📄 Mostrar Bitácoras", "💬 Comentarios"])


# ---------------- helper: preparar df con tipos
def _prep(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    df["_fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    return df


def construir_bloque_indicadores(fdf, fdf_act, etiqueta, horas_acum, horas_tot):
    """
    Dado un subconjunto ya filtrado de atenciones (fdf) y de actividades
    extra (fdf_act), arma el dict de KPIs y la lista de (título, figura)
    para el PDF. Se usa tanto para el reporte de un periodo como para
    cada semana del reporte compilado 1-10, así los dos salen idénticos.
    """
    total = len(fdf)
    horas_mant = fdf["duracion_min"].dropna().sum() / 60.0 if not fdf.empty else 0.0
    horas_extra = fdf_act["horas"].dropna().sum() if not fdf_act.empty else 0.0
    prog = (horas_acum / horas_tot * 100) if horas_tot else 0

    kpis = {
        "Equipos atendidos": str(total),
        "Horas acumuladas": f"{horas_acum} h",
        "Horas mantenimiento": f"{horas_mant:.1f} h",
        "Otras actividades": f"{horas_extra:.1f} h",
        "Progreso": f"{prog:.0f}%",
    }

    figuras = []

    # 1) Distribución de tiempo por tipo de actividad
    extra_por_tipo = (fdf_act.groupby("tipo_actividad")["horas"].sum()
                      if not fdf_act.empty else pd.Series(dtype=float))
    dist = pd.concat([
        pd.Series({"Mantenimiento de equipos": horas_mant}),
        extra_por_tipo,
    ]).rename_axis("Tipo de actividad").reset_index(name="Horas")
    dist = dist[dist["Horas"] > 0].sort_values("Horas", ascending=True)
    if not dist.empty:
        PAL = ["#1F4E78", "#2E8B57", "#C0392B", "#8E44AD", "#D68910"]
        fdist = px.bar(dist, x="Horas", y="Tipo de actividad", orientation="h",
                       text="Horas", color="Tipo de actividad",
                       color_discrete_sequence=PAL)
        fdist.update_traces(texttemplate="%{text:.1f} h", textposition="outside")
        fdist.update_layout(showlegend=False, yaxis_title="", xaxis_title="Horas")
        figuras.append(("Distribucion de tiempo", fdist))

    if not fdf.empty:
        # 2) Equipos por área
        por_area = (fdf.groupby("area").size().rename_axis("Área")
                    .reset_index(name="Equipos"))
        por_area = por_area[por_area["Equipos"] > 0]
        if not por_area.empty:
            fa = px.bar(por_area, x="Área", y="Equipos", text="Equipos")
            fa.update_layout(xaxis_tickangle=-40)
            figuras.append(("Equipos por area", fa))

        # 3) Tipo de mantenimiento
        por_tipo = (fdf.groupby("tipo").size().rename_axis("Tipo")
                    .reset_index(name="Cantidad"))
        if not por_tipo.empty:
            ft = px.pie(por_tipo, names="Tipo", values="Cantidad", hole=0.4)
            figuras.append(("Tipo de mantenimiento", ft))

        # 4) Estado de resolución
        por_res = (fdf.groupby("resuelto").size()
                   .reindex(RESUELTO, fill_value=0)
                   .rename_axis("Estado").reset_index(name="Cantidad"))
        por_res = por_res[por_res["Cantidad"] > 0]
        if not por_res.empty:
            fr = px.pie(por_res, names="Estado", values="Cantidad", hole=0.4)
            figuras.append(("Estado de resolucion", fr))

        # 5) y 6) Solo para el reporte de toda la práctica: tendencia de
        # horas acumuladas (real vs meta) y evolución del tipo por semana.
        if etiqueta == "Toda la práctica":
            semanas_lst = cal.lista_semanas()
            meta_sem = (horas_tot / len(semanas_lst)) if semanas_lst else 0
            hoy_ref = hoy_hn()
            filas_t = []
            for n in semanas_lst:
                vsem = cal.viernes_de_semana(n)
                filas_t.append({
                    "Semana": f"Sem {n}",
                    "Meta": round(meta_sem * n),
                    "Real": (cal.horas_hasta(vsem) if vsem <= hoy_ref else None)})
            tend = pd.DataFrame(filas_t)
            ftend = go.Figure()
            ftend.add_trace(go.Scatter(
                x=tend["Semana"], y=tend["Meta"], name="Meta (400 h)",
                mode="lines", line=dict(color="#C7D2E8", dash="dash", width=2)))
            ftend.add_trace(go.Scatter(
                x=tend["Semana"], y=tend["Real"], name="Avance real",
                mode="lines+markers", line=dict(color="#0EA5A5", width=3),
                marker=dict(size=7, color="#0EA5A5"), connectgaps=False))
            ftend.update_layout(yaxis_title="Horas")
            figuras.append(("Tendencia de horas acumuladas", ftend))

            evol = (fdf.dropna(subset=["semana"])
                    .groupby(["semana", "tipo"]).size()
                    .reset_index(name="Cantidad"))
            if not evol.empty:
                evol["Semana"] = evol["semana"].apply(lambda n: f"Sem {n}")
                fev = px.bar(evol, x="Semana", y="Cantidad", color="tipo",
                             color_discrete_sequence=["#2563EB", "#0EA5A5",
                                                      "#7C3AED", "#F59E0B",
                                                      "#EF4444", "#16A34A"])
                fev.update_layout(barmode="stack", yaxis_title="Equipos",
                                  xaxis_title="")
                figuras.append(("Evolucion del tipo de mantenimiento", fev))

    return kpis, figuras


# =============================================================== TAB: Nueva atención
with tab_input:
    st.subheader("Registrar una atención a equipo")

    # ---- Carga desde foto con IA (pre-rellena el formulario) ----
    # Los valores extraídos se guardan en session_state y el formulario de
    # abajo los usa como valores por defecto, para que puedas revisarlos y
    # editarlos antes de guardar.
    _AI = st.session_state.setdefault("ai_pre", {})

    with st.expander("📷 Cargar desde foto (IA)"):
        st.caption("Al subir una foto o PDF del reporte, la IA extrae los "
                   "datos y rellena el formulario para revisarlos antes de "
                   "guardar.")
        ai_file = st.file_uploader(
            "Foto o PDF del reporte", type=["jpg", "jpeg", "png", "webp", "pdf"],
            key="ai_upload", accept_multiple_files=False)
        if st.button("Extraer datos con IA", key="btn_ai",
                     disabled=ai_file is None):
            api_key = st.secrets.get("GEMINI_API_KEY", "")
            if not api_key:
                st.error("Falta la clave GEMINI_API_KEY en los secretos de la "
                         "app. Agrégala en .streamlit/secrets.toml (es la misma "
                         "de ServiDox).")
            else:
                try:
                    with st.spinner("Leyendo el reporte con IA..."):
                        contenido = ai_file.getvalue()
                        datos, mime = ai_extract.preparar_archivo(
                            ai_file.name, contenido)
                        extraido = ai_extract.extraer_de_imagen(
                            datos, mime, api_key)
                    st.session_state["ai_pre"] = extraido
                    st.success("Datos extraídos. Quedan cargados en el "
                               "formulario para revisión y ajuste antes de "
                               "guardar.")
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo extraer: {e}")
        if _AI:
            if st.button("Limpiar datos extraídos", key="btn_ai_clear"):
                st.session_state["ai_pre"] = {}
                st.rerun()

    # Helpers para valores por defecto: si la IA extrajo algo, se usa; si no,
    # el valor normal.
    def _pre_txt(campo):
        return _AI.get(campo, "") or ""

    def _pre_idx(campo, opciones, default=0):
        val = _AI.get(campo, "")
        return opciones.index(val) if val in opciones else default

    def _pre_fecha():
        val = _AI.get("fecha", "")
        if val:
            try:
                return date.fromisoformat(val)
            except ValueError:
                pass
        return hoy_hn()

    def _pre_hora(campo, default):
        val = _AI.get(campo, "")
        if val:
            try:
                h, m = val.split(":")[:2]
                return time(int(h), int(m))
            except (ValueError, IndexError):
                pass
        return default

    with st.form("nueva", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            f_fecha = st.date_input("Fecha", value=_pre_fecha())
            f_area = st.selectbox("Área", AREAS,
                                  index=_pre_idx("area", AREAS))
            f_tipo = st.selectbox("Tipo de mantenimiento", TIPOS,
                                  index=_pre_idx("tipo", TIPOS))
        with c2:
            f_equipo = st.text_input("Equipo", value=_pre_txt("equipo"),
                                     placeholder="Lámpara cielítica")
            f_marca = st.text_input("Marca", value=_pre_txt("marca"),
                                    placeholder="Dräger")
            f_modelo = st.text_input("Modelo", value=_pre_txt("modelo"),
                                     placeholder="Polaris 100")
        with c3:
            f_serie = st.text_input("Serie No.", value=_pre_txt("serie"),
                                    placeholder="ABC123")
            f_ini = st.time_input("Hora de inicio",
                                  value=_pre_hora("hora_inicio", time(8, 0)))
            f_fin = st.time_input("Hora de finalización",
                                  value=_pre_hora("hora_fin", time(9, 0)))

        # Semana y Día EDITABLES. Se sugieren desde la fecha, pero podés
        # corregirlos (p.ej. si cargás hoy un registro que era de ayer).
        _sem_sug = cal.semana_de_fecha(f_fecha) or 1
        _dia_sug = cal.dia_de_semana_num(f_fecha) or 1
        s1, s2, _s3 = st.columns([1, 1, 2])
        with s1:
            f_semana = st.selectbox(
                "Semana", cal.lista_semanas(),
                index=cal.lista_semanas().index(_sem_sug),
                help="Corrige la semana si el registro corresponde a otra.")
        with s2:
            f_dia = st.selectbox(
                "Día", [1, 2, 3, 4, 5], index=_dia_sug - 1,
                format_func=lambda d: f"Día {d}",
                help="Corrige el día si lo cargas con retraso.")

        f_prob = st.text_area("Problema identificado",
                              value=_pre_txt("problema"), height=80)
        f_sol = st.text_area("Solución sugerida / trabajo realizado",
                             value=_pre_txt("solucion"), height=80)

        cc1, cc2 = st.columns([1, 3])
        with cc1:
            f_res = st.selectbox("¿Resuelto?", RESUELTO,
                                 index=_pre_idx("resuelto", RESUELTO))
        f_imp = st.text_area("Impacto esperado o beneficio real",
                             value=_pre_txt("impacto"), height=70)
        f_obs = st.text_area("Observaciones",
                             value=_pre_txt("observaciones"), height=70)

        # Hasta 4 imágenes JPEG que se mostrarán en "Mostrar Bitácoras"
        f_imgs = st.file_uploader(
            "Evidencia fotográfica (hasta 4 imágenes JPEG)",
            type=["jpg", "jpeg", "png"], accept_multiple_files=True,
            key="up_imgs",
            help="Arrastra o selecciona hasta 4 fotos. Se comprimen automáticamente.")

        enviado = st.form_submit_button("Guardar atención", type="primary")
        if enviado:
            if not f_equipo.strip():
                st.error("El campo 'Equipo' es obligatorio.")
            else:
                # Procesar imágenes (máx 4)
                imgs_b64 = []
                if f_imgs:
                    for archivo in f_imgs[:4]:
                        b64 = _img_a_base64(archivo)
                        if b64:
                            imgs_b64.append(b64)
                    if len(f_imgs) > 4:
                        st.warning("Solo se guardaron las primeras 4 imágenes.")
                while len(imgs_b64) < 4:
                    imgs_b64.append(None)

                if cal.semana_de_fecha(f_fecha) is None:
                    st.warning("⚠️ La fecha está fuera del período de práctica "
                               "(20 jul – 25 sep 2026), pero se guardará con la "
                               f"Semana {f_semana} y Día {f_dia} que elegiste.")
                insert_atencion({
                    "fecha": f_fecha.isoformat(),
                    "semana": int(f_semana), "dia": int(f_dia),
                    "hora_inicio": f_ini.strftime("%H:%M"),
                    "hora_fin": f_fin.strftime("%H:%M"),
                    "area": f_area, "equipo": f_equipo.strip(),
                    "marca": f_marca.strip(), "modelo": f_modelo.strip(),
                    "serie": f_serie.strip(), "tipo": f_tipo,
                    "problema": f_prob.strip(), "solucion": f_sol.strip(),
                    "resuelto": f_res, "impacto": f_imp.strip(),
                    "observaciones": f_obs.strip(),
                    "img1": imgs_b64[0], "img2": imgs_b64[1],
                    "img3": imgs_b64[2], "img4": imgs_b64[3],
                })
                n_fotos = sum(1 for x in imgs_b64 if x)
                extra = f" con {n_fotos} imagen(es)" if n_fotos else ""
                st.success(f"✅ Atención registrada (Semana {f_semana}, "
                           f"Día {f_dia}{extra}).")
                # Limpiar los datos pre-rellenados por la IA para el próximo
                # registro.
                if st.session_state.get("ai_pre"):
                    st.session_state["ai_pre"] = {}

    # -------------------------------------------- Otras actividades (sin equipo)
    st.divider()
    st.subheader("🕒 Otra actividad (no mantenimiento de equipo)")
    st.caption("Para visitas, infografías, protocolos, reuniones u otras "
               "actividades de la práctica que no son la atención puntual "
               "de un equipo: registra solo el tiempo invertido.")

    # El selectbox de tipo va FUERA del form para poder mostrar el campo
    # libre de "¿Cuál actividad?" de inmediato cuando elige "Otro" (dentro
    # de un st.form los widgets no reaccionan hasta enviar).
    a_tipo_sel = st.selectbox("Actividad", TIPOS_ACTIVIDAD_EXTRA, key="af_tipo")
    a_otro = ""
    if a_tipo_sel == "Otro":
        a_otro = st.text_input(
            "¿Cuál actividad? (si coincide con una ya registrada, sus horas "
            "se suman juntas automáticamente)",
            placeholder="Ej. Auditoría de inventario", key="af_otro")

    with st.form("nueva_actividad", clear_on_submit=True):
        ac1, ac2, ac3 = st.columns(3)
        with ac1:
            a_fecha = st.date_input("Fecha", value=hoy_hn(), key="af_fecha")
        with ac2:
            _sem_sug2 = cal.semana_de_fecha(a_fecha) or 1
            _dia_sug2 = cal.dia_de_semana_num(a_fecha) or 1
            a_semana = st.selectbox(
                "Semana", cal.lista_semanas(),
                index=cal.lista_semanas().index(_sem_sug2), key="af_sem")
            a_dia = st.selectbox("Día", [1, 2, 3, 4, 5], index=_dia_sug2 - 1,
                                 format_func=lambda d: f"Día {d}", key="af_dia")
        with ac3:
            a_horas = st.number_input(
                "Tiempo invertido (horas)", min_value=0.0, max_value=12.0,
                step=0.5, value=1.0, key="af_horas")

        a_desc = st.text_input(
            "Descripción (opcional)",
            placeholder="Ej. Visita a CIPS de Occidente", key="af_desc")

        env2 = st.form_submit_button("Guardar actividad", type="primary")
        if env2:
            if a_tipo_sel == "Otro" and not a_otro.strip():
                st.error("Escribe el nombre de la actividad en '¿Cuál actividad?'.")
            else:
                tipo_final = (normalizar_tipo_actividad(a_otro)
                             if a_tipo_sel == "Otro" else a_tipo_sel)
                insert_actividad_extra({
                    "fecha": a_fecha.isoformat(), "semana": int(a_semana),
                    "dia": int(a_dia), "tipo_actividad": tipo_final,
                    "descripcion": a_desc.strip(), "horas": float(a_horas),
                })
                st.success(f"✅ Actividad registrada ({a_horas} h, "
                           f"Semana {a_semana}, Día {a_dia}).")

    with st.expander("Ver / editar / eliminar actividades registradas"):
        dfa = load_actividades_extra()
        if dfa.empty:
            st.caption("Aún no hay actividades registradas.")
        else:
            dfa_show = dfa.sort_values(
                ["semana", "dia", "fecha"], na_position="last").reset_index(drop=True)
            show_a = dfa_show.rename(columns={
                "id": "ID", "fecha": "Fecha", "semana": "Semana", "dia": "Día",
                "tipo_actividad": "Actividad", "descripcion": "Descripción",
                "horas": "Horas"})
            st.caption("Editar cualquier celda y presionar 'Guardar cambios'. "
                       "Para borrar una fila, seleccionarla con el check de la "
                       "izquierda y presionar la papelera. Si se escribe en "
                       "'Actividad' un nombre igual a uno ya existente (aunque "
                       "cambien mayúsculas o espacios), se fusiona con ese y "
                       "las horas se suman en el mismo grupo del gráfico.")
            edited_act = st.data_editor(
                show_a, use_container_width=True, hide_index=True,
                num_rows="dynamic", key="editor_actividades",
                column_config={
                    "ID": st.column_config.NumberColumn("ID", disabled=True),
                    "Horas": st.column_config.NumberColumn(
                        "Horas", min_value=0.0, step=0.5, format="%.1f"),
                    "Semana": st.column_config.NumberColumn(
                        "Semana", min_value=1, step=1),
                    "Día": st.column_config.NumberColumn(
                        "Día", min_value=1, max_value=5, step=1),
                })
            if st.button("💾 Guardar cambios", key="save_act"):
                inv_a = {"ID": "id", "Fecha": "fecha", "Semana": "semana",
                         "Día": "dia", "Actividad": "tipo_actividad",
                         "Descripción": "descripcion", "Horas": "horas"}
                update_actividades_extra_from_df(edited_act.rename(columns=inv_a))
                st.success("✅ Cambios guardados.")
                st.rerun()


# =============================================================== TAB: Cargar bitácora
with tab_carga:
    st.subheader("Cargar bitácora institucional (Matriz de impacto)")
    st.caption("Permite subir un Excel de la Matriz de impacto (UNITEC). BitaLogs "
               "detecta semana, día, equipo, problema, solución, ¿resuelto?, "
               "impacto y observaciones. Luego asignas fecha, horas, área y "
               "tipo antes de guardar.")

    up = st.file_uploader("Archivo .xlsx", type=["xlsx"], key="up_bitacora")
    if up is not None:
        try:
            regs, avisos = leer_matriz(up)
        except Exception as e:
            st.error(f"No se pudo leer el archivo: {e}")
            regs, avisos = [], []

        for a in avisos:
            st.warning(a)

        if regs:
            st.success(f"Se detectaron **{len(regs)}** registro(s).")
            prev = pd.DataFrame(regs)

            # Sugerir fecha a partir de semana+día del formato
            def _fecha_sug(row):
                if row["semana"] and row["dia"]:
                    try:
                        return cal.fecha_de_semana_dia(
                            int(row["semana"]), int(row["dia"])).isoformat()
                    except Exception:
                        return ""
                return ""

            prev.insert(0, "fecha", prev.apply(_fecha_sug, axis=1))
            prev["area"] = ""
            prev["tipo"] = "Correctivo"
            prev["hora_inicio"] = "08:00"
            prev["hora_fin"] = "09:00"
            prev["marca"] = ""; prev["modelo"] = ""; prev["serie"] = ""

            st.markdown("**Revisa y completa antes de guardar** "
                        "(fecha, área, tipo y horas):")
            edit = st.data_editor(
                prev, use_container_width=True, hide_index=True,
                num_rows="dynamic", key="editor_carga",
                column_config={
                    "area": st.column_config.SelectboxColumn("Área", options=AREAS),
                    "tipo": st.column_config.SelectboxColumn(
                        "Tipo", options=TIPOS),
                    "resuelto": st.column_config.SelectboxColumn(
                        "¿Resuelto?", options=RESUELTO),
                },
            )
            if st.button("💾 Guardar todos en BitaLogs", type="primary",
                         key="save_carga"):
                nuevos = []
                for _, r in edit.iterrows():
                    fecha = str(r.get("fecha", "")).strip()[:10]
                    if not str(r.get("equipo", "")).strip():
                        continue
                    try:
                        fd = datetime.strptime(fecha, "%Y-%m-%d").date()
                        sem = cal.semana_de_fecha(fd)
                        dia = cal.dia_de_semana_num(fd)
                    except ValueError:
                        sem = r.get("semana"); dia = r.get("dia")
                    nuevos.append({
                        "fecha": fecha or hoy_hn().isoformat(),
                        "semana": sem, "dia": dia,
                        "hora_inicio": str(r.get("hora_inicio", "") or ""),
                        "hora_fin": str(r.get("hora_fin", "") or ""),
                        "area": r.get("area", ""), "equipo": r.get("equipo", ""),
                        "marca": r.get("marca", ""), "modelo": r.get("modelo", ""),
                        "serie": r.get("serie", ""), "tipo": r.get("tipo", "Otro"),
                        "problema": r.get("problema", ""),
                        "solucion": r.get("solucion", ""),
                        "resuelto": r.get("resuelto", "No"),
                        "impacto": r.get("impacto", ""),
                        "observaciones": r.get("observaciones", ""),
                    })
                insert_muchas(nuevos)
                st.success(f"✅ {len(nuevos)} atención(es) guardada(s).")
                st.rerun()


# =============================================================== TAB: Dashboard
with tab_dash:
    df = _prep(load_atenciones())
    if df.empty:
        st.info("Aún no hay atenciones. Se registran en '➕ Nuevo registro' "
                "o se carga una bitácora en '📥 Cargar bitácora'.")
    else:
        with st.sidebar:
            st.header("Filtros")
            modo = st.radio("Ver por:", ["Toda la práctica", "Semana", "Día"],
                            key="modo_filtro")
            sem_sel = None
            dia_fecha = None
            if modo == "Semana":
                sem_sel = st.selectbox(
                    "Semana", cal.lista_semanas(),
                    format_func=lambda n: cal.etiqueta_semana(n))
            elif modo == "Día":
                sem_pick = st.selectbox(
                    "Semana", cal.lista_semanas(),
                    format_func=lambda n: f"Semana {n}", key="dia_sem")
                dnum = st.selectbox("Día", [1, 2, 3, 4, 5],
                                    format_func=lambda d: cal.etiqueta_dia(
                                        cal.fecha_de_semana_dia(sem_pick, d)))
                dia_fecha = cal.fecha_de_semana_dia(sem_pick, dnum)

            st.divider()
            fa = st.multiselect("Área", AREAS)
            ft = st.multiselect("Tipo", TIPOS)

        fdf = df.copy()
        if modo == "Semana" and sem_sel:
            fdf = fdf[fdf["semana"] == sem_sel]
        elif modo == "Día" and dia_fecha is not None:
            fdf = fdf[fdf["_fecha"].dt.date == dia_fecha]
        if fa:
            fdf = fdf[fdf["area"].isin(fa)]
        if ft:
            fdf = fdf[fdf["tipo"].isin(ft)]

        # ---- Horas (se calculan primero para el hero) ----
        hoy = hoy_hn()
        ref = hoy
        if modo == "Semana" and sem_sel:
            ref = cal.viernes_de_semana(sem_sel)
        elif modo == "Día" and dia_fecha is not None:
            ref = dia_fecha
        horas_acum = cal.horas_hasta(ref)
        horas_tot = cal.horas_totales_practica()
        sem_actual = cal.semana_de_fecha(ref) or (
            cal.TOTAL_SEMANAS if ref >= cal.fin_practica() else 1)
        progreso = horas_acum / horas_tot * 100 if horas_tot else 0

        # Actividades extra (para el KPI de "otras actividades")
        _dfa_hero = load_actividades_extra()
        _dfa_hero["_fecha"] = pd.to_datetime(_dfa_hero.get("fecha"),
                                             errors="coerce")
        _fa = _dfa_hero.copy()
        if modo == "Semana" and sem_sel:
            _fa = _fa[_fa["semana"] == sem_sel]
        elif modo == "Día" and dia_fecha is not None:
            _fa = _fa[_fa["_fecha"].dt.date == dia_fecha]
        horas_mant_hero = (fdf["duracion_min"].dropna().sum() / 60.0
                           if not fdf.empty else 0.0)
        horas_extra_hero = (_fa["horas"].dropna().sum()
                            if not _fa.empty else 0.0)

        # ---- Hero ----
        if modo == "Semana" and sem_sel:
            _periodo_txt = f"Semana {sem_sel}"
        elif modo == "Día" and dia_fecha is not None:
            _periodo_txt = cal.etiqueta_dia(dia_fecha)
        else:
            _periodo_txt = "Toda la práctica"
        ui.hero(
            subtitulo="Práctica Profesional · Ingeniería Biomédica · UNITEC",
            horas_acum=horas_acum, horas_tot=horas_tot,
            periodo_txt=_periodo_txt, progreso=progreso)

        # ---- KPIs principales ----
        total = len(fdf)
        equipos_unicos = fdf["equipo"].str.strip().str.lower().nunique()
        dur_prom = fdf["duracion_min"].dropna().mean()
        dias_activos = fdf["_fecha"].dt.date.nunique()
        eq_por_dia = (total / dias_activos) if dias_activos else 0

        ui.fila_kpis([
            {"valor": str(total), "label": "Equipos atendidos",
             "icono": "equipos"},
            {"valor": f"{horas_acum}", "unidad": "h",
             "label": "Horas acumuladas", "icono": "reloj"},
            {"valor": f"{horas_mant_hero:.1f}", "unidad": "h",
             "label": "Mantenimiento", "icono": "llave"},
            {"valor": f"{horas_extra_hero:.1f}", "unidad": "h",
             "label": "Otras actividades", "icono": "lista"},
            {"valor": f"{progreso:.0f}", "unidad": "%",
             "label": "Progreso", "icono": "tendencia"},
        ])


        # ---- Distribución de tiempo por tipo de actividad ----
        # Combina las horas de mantenimiento de equipo (calculadas de
        # hora_inicio/hora_fin en 'atenciones') con las horas de otras
        # actividades (visitas, infografías, etc.) cargadas en
        # 'actividades_extra', respetando el mismo filtro de Semana/
        # Día/Toda la práctica que el resto del Dashboard.
        ui.encabezado_seccion("Distribución de tiempo por tipo de actividad",
                              ui.PALETA[1])

        df_act = load_actividades_extra()
        df_act["_fecha"] = pd.to_datetime(df_act.get("fecha"), errors="coerce")
        fdf_act = df_act.copy()
        if modo == "Semana" and sem_sel:
            fdf_act = fdf_act[fdf_act["semana"] == sem_sel]
        elif modo == "Día" and dia_fecha is not None:
            fdf_act = fdf_act[fdf_act["_fecha"].dt.date == dia_fecha]

        horas_mant = fdf["duracion_min"].dropna().sum() / 60.0
        horas_extra_tot = fdf_act["horas"].dropna().sum() if not fdf_act.empty else 0.0
        extra_por_tipo = (fdf_act.groupby("tipo_actividad")["horas"].sum()
                          if not fdf_act.empty else pd.Series(dtype=float))

        dist = pd.concat([
            pd.Series({"Mantenimiento de equipos": horas_mant}),
            extra_por_tipo,
        ]).rename_axis("Tipo de actividad").reset_index(name="Horas")
        dist = dist[dist["Horas"] > 0].sort_values("Horas", ascending=True)

        if dist.empty:
            st.info("Aún no hay horas registradas (ni de mantenimiento ni de "
                    "otras actividades) para este filtro.")
            fig_dist = None
        else:
            fig_dist = px.bar(dist, x="Horas", y="Tipo de actividad",
                              orientation="h", text="Horas",
                              color="Tipo de actividad",
                              color_discrete_sequence=ui.PALETA)
            fig_dist.update_traces(
                texttemplate="%{text:.1f} h", textposition="outside",
                cliponaxis=False)
            for _, fila_d in dist.iterrows():
                fig_dist.add_annotation(
                    x=0, y=fila_d["Tipo de actividad"],
                    text=f"<b>{fila_d['Tipo de actividad']}</b>",
                    showarrow=False, xanchor="left", yanchor="bottom",
                    yshift=16, xshift=-2,
                    font=dict(size=13, color="#1B2436"))
            altura_dist = max(360, 62 * len(dist) + 60)
            ui.estilizar_figura(fig_dist, altura=altura_dist, leyenda=False)
            fig_dist.update_layout(
                yaxis=dict(showticklabels=False, title=""),
                xaxis=dict(title="Horas"),
                margin=dict(l=8, r=45, t=14, b=40),
                bargap=0.55,
                uniformtext=dict(mode="hide", minsize=10))
            st.plotly_chart(fig_dist, use_container_width=True)

        # ---- Gráficos de rendimiento ----
        g1, g2 = st.columns(2)
        with g1:
            por_area = (fdf.groupby("area").size()
                        .reindex(AREAS, fill_value=0)
                        .rename_axis("Área").reset_index(name="Equipos"))
            por_area = por_area[por_area["Equipos"] > 0]
            ui.encabezado_seccion("Equipos por área", ui.PALETA[0])
            if por_area.empty:
                st.info("Aún no hay atenciones con área asignada para mostrar.")
            else:
                fig = px.bar(por_area, x="Área", y="Equipos",
                             color_discrete_sequence=[ui.PALETA[0]])
                fig.update_layout(xaxis_tickangle=-40)
                ui.estilizar_figura(fig, altura=290, leyenda=False)
                st.plotly_chart(fig, use_container_width=True)
        with g2:
            por_tipo = (fdf.groupby("tipo").size()
                        .rename_axis("Tipo").reset_index(name="Cantidad"))
            ui.encabezado_seccion("Tipo de mantenimiento", ui.PALETA[2])
            if len(por_tipo) == 1:
                # Una sola categoría: tarjeta clara en vez de donut "vacío".
                ui.tarjeta_categoria_unica(por_tipo.iloc[0]["Tipo"],
                                           color=ui.PALETA[2])
            elif not por_tipo.empty:
                fig2 = px.pie(por_tipo, names="Tipo", values="Cantidad",
                              hole=0.55, color_discrete_sequence=ui.PALETA)
                fig2.update_traces(textposition="inside", textinfo="percent")
                ui.estilizar_figura(fig2, altura=290, leyenda=True)
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("Sin datos de mantenimiento para este filtro.")

        g3, g4 = st.columns(2)
        with g3:
            por_res = (fdf.groupby("resuelto").size()
                       .reindex(RESUELTO, fill_value=0)
                       .rename_axis("¿Resuelto?").reset_index(name="Cantidad"))
            ui.encabezado_seccion("Estado de resolución", ui.PALETA[5])
            fig3 = px.pie(por_res, names="¿Resuelto?", values="Cantidad",
                          hole=0.55,
                          color="¿Resuelto?",
                          color_discrete_map={"Sí": ui.PALETA[5],
                                              "Parcial": ui.PALETA[3],
                                              "No": ui.PALETA[4]})
            fig3.update_traces(textposition="inside", textinfo="percent")
            ui.estilizar_figura(fig3, altura=290, leyenda=True)
            st.plotly_chart(fig3, use_container_width=True)
        with g4:
            dd = fdf.dropna(subset=["duracion_min"])
            ui.encabezado_seccion("Duración de las atenciones", ui.PALETA[6])
            if not dd.empty:
                # Se agrupa la duración en rangos y se grafica como barras
                # (no histograma) para que respeten las esquinas redondeadas
                # del resto del dashboard.
                bins = [0, 15, 30, 45, 60, 90, 120, 9999]
                etiquetas = ["0-15", "15-30", "30-45", "45-60",
                             "60-90", "90-120", "120+"]
                dd = dd.copy()
                dd["rango"] = pd.cut(dd["duracion_min"], bins=bins,
                                     labels=etiquetas, right=False)
                por_dur = (dd.groupby("rango", observed=False).size()
                           .reindex(etiquetas, fill_value=0)
                           .rename_axis("Rango").reset_index(name="Atenciones"))
                fig4 = px.bar(por_dur, x="Rango", y="Atenciones",
                              text="Atenciones",
                              color_discrete_sequence=[ui.PALETA[6]])
                fig4.update_traces(textposition="outside", cliponaxis=False)
                _max4 = por_dur["Atenciones"].max()
                fig4.update_layout(xaxis_title="Minutos",
                                   yaxis_title="Atenciones",
                                   yaxis_range=[0, (_max4 * 1.18) if _max4 else 1])
                ui.estilizar_figura(fig4, altura=290, leyenda=False)
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info("Sin datos de duración para graficar.")

        # ---- Equipos por semana ----
        if modo == "Toda la práctica":
            por_sem = (fdf.dropna(subset=["semana"]).groupby("semana").size()
                       .reindex(cal.lista_semanas(), fill_value=0)
                       .rename_axis("Semana").reset_index(name="Equipos"))
            por_sem["Semana"] = por_sem["Semana"].apply(lambda n: f"Sem {n}")
            ui.encabezado_seccion("Equipos atendidos por semana", ui.PALETA[1])
            fig5 = px.bar(por_sem, x="Semana", y="Equipos", text="Equipos",
                          color_discrete_sequence=[ui.PALETA[1]])
            fig5.update_traces(textposition="outside", cliponaxis=False)
            _max5 = por_sem["Equipos"].max()
            fig5.update_layout(
                yaxis_range=[0, (_max5 * 1.18) if _max5 else 1])
            ui.estilizar_figura(fig5, altura=280, leyenda=False)
            st.plotly_chart(fig5, use_container_width=True)

            # ---- Tendencia de horas acumuladas: real vs meta ----
            # Responde a la sugerencia de proyectar el cumplimiento de las
            # 400 h: la meta ideal (lineal) contra el avance real semana
            # a semana.
            semanas_lst = cal.lista_semanas()
            horas_tot_pract = cal.horas_totales_practica()
            meta_por_sem = horas_tot_pract / len(semanas_lst)
            hoy_ref = hoy_hn()
            filas_tend = []
            for n in semanas_lst:
                vsem = cal.viernes_de_semana(n)
                real_val = (cal.horas_hasta(vsem) if vsem <= hoy_ref
                            else None)
                filas_tend.append({
                    "Semana": f"Sem {n}",
                    "Meta": round(meta_por_sem * n),
                    "Real": real_val})
            tend = pd.DataFrame(filas_tend)
            ui.encabezado_seccion("Tendencia de horas acumuladas",
                                  ui.PALETA[1])
            fig_t = go.Figure()
            fig_t.add_trace(go.Scatter(
                x=tend["Semana"], y=tend["Meta"], name="Meta (400 h)",
                mode="lines", line=dict(color="#C7D2E8", dash="dash", width=2)))
            fig_t.add_trace(go.Scatter(
                x=tend["Semana"], y=tend["Real"], name="Avance real",
                mode="lines+markers",
                line=dict(color=ui.PALETA[1], width=3),
                marker=dict(size=7, color=ui.PALETA[1]),
                connectgaps=False))
            fig_t.update_layout(yaxis_title="Horas")
            ui.estilizar_figura(fig_t, altura=300, leyenda=True)
            st.plotly_chart(fig_t, use_container_width=True)

            # ---- Evolución del tipo de mantenimiento por semana ----
            # Barras apiladas: muestra si se está pasando de correctivo
            # (reactivo) a preventivo con el paso de las semanas.
            evol = (fdf.dropna(subset=["semana"])
                    .groupby(["semana", "tipo"]).size()
                    .reset_index(name="Cantidad"))
            if not evol.empty:
                evol["Semana"] = evol["semana"].apply(lambda n: f"Sem {n}")
                evol = evol.sort_values("semana")
                ui.encabezado_seccion(
                    "Evolución del tipo de mantenimiento", ui.PALETA[2])
                fig_ev = px.bar(
                    evol, x="Semana", y="Cantidad", color="tipo",
                    color_discrete_sequence=ui.PALETA,
                    labels={"tipo": "Tipo", "Cantidad": "Equipos"})
                fig_ev.update_layout(barmode="stack", yaxis_title="Equipos",
                                     xaxis_title="")
                ui.estilizar_figura(fig_ev, altura=300, leyenda=True)
                st.plotly_chart(fig_ev, use_container_width=True)

        # ---- Reporte PDF (dashboard, imprimible) ----
        st.divider()
        if modo == "Semana" and sem_sel:
            etiqueta_periodo = f"Semana {sem_sel}"
        elif modo == "Día" and dia_fecha is not None:
            etiqueta_periodo = cal.etiqueta_dia(dia_fecha)
        else:
            etiqueta_periodo = "Toda la práctica"

        st.subheader(f"📄 Reporte PDF: {etiqueta_periodo}")

        cpdf1, cpdf2 = st.columns(2)

        with cpdf1:
            if st.button("📄 PDF global de práctica", key="gen_pdf_dash",
                         use_container_width=True):
                with st.spinner("Generando PDF..."):
                    kpis_pdf, figuras_pdf = construir_bloque_indicadores(
                        fdf, fdf_act, etiqueta_periodo, horas_acum, horas_tot)
                    # Comentarios sensibles al filtro: si es una semana,
                    # solo los de esa semana; si es toda la práctica, todos.
                    _dcom = load_comentarios()
                    if modo == "Semana" and sem_sel:
                        _dcom = _dcom[_dcom["semana"] == sem_sel]
                    coment_pdf = _dcom.sort_values(
                        ["semana", "fecha_registro"]).rename(columns={
                            "fecha_registro": "fecha"}).to_dict("records")
                    pdf_bytes = construir_pdf(
                        titulo="BitaLogs - Reporte de Indicadores",
                        subtitulo=etiqueta_periodo,
                        kpis=kpis_pdf, figuras=figuras_pdf,
                        progreso=horas_acum / horas_tot * 100 if horas_tot else 0,
                        horas_acum=horas_acum, horas_tot=horas_tot,
                        comentarios=coment_pdf)
                st.download_button(
                    "📥 Descargar PDF", data=pdf_bytes,
                    file_name=f"BitaLogs_{etiqueta_periodo.replace(' ', '_')}.pdf",
                    mime="application/pdf", key="dl_pdf_dash",
                    use_container_width=True)

        with cpdf2:
            if st.button("📚 PDF semana 1-10", key="gen_pdf_multi",
                         use_container_width=True):
                with st.spinner("Generando PDF de todas las semanas..."):
                    df_full = _prep(load_atenciones())
                    df_act_full = load_actividades_extra()
                    df_act_full["_fecha"] = pd.to_datetime(
                        df_act_full.get("fecha"), errors="coerce")
                    coment_full = load_comentarios()
                    bloques = []
                    for s in cal.lista_semanas():
                        sub = df_full[df_full["semana"] == s] if not df_full.empty else df_full
                        sub_act = (df_act_full[df_act_full["semana"] == s]
                                   if not df_act_full.empty else df_act_full)
                        if sub.empty and sub_act.empty:
                            continue  # saltar semanas sin nada registrado
                        h_acum = cal.horas_hasta(cal.viernes_de_semana(s))
                        kpis_s, figs_s = construir_bloque_indicadores(
                            sub, sub_act, f"Semana {s}", h_acum, horas_tot)
                        # Comentarios de esta semana
                        com_s = (coment_full[coment_full["semana"] == s]
                                 if not coment_full.empty else coment_full)
                        com_s = com_s.rename(columns={
                            "fecha_registro": "fecha"}).to_dict("records") \
                            if not com_s.empty else []
                        bloques.append({
                            "subtitulo": f"Semana {s}",
                            "kpis": kpis_s, "figuras": figs_s,
                            "progreso": h_acum / horas_tot * 100 if horas_tot else 0,
                            "horas_acum": h_acum, "horas_tot": horas_tot,
                            "comentarios": com_s})

                    if not bloques:
                        st.warning("No hay datos registrados en ninguna semana.")
                    else:
                        pdf_multi = construir_pdf_multi(
                            "BitaLogs - Reporte Semanal", bloques)
                        st.download_button(
                            "📥 Descargar PDF (10 semanas)", data=pdf_multi,
                            file_name="BitaLogs_Reporte_Semanas_1-10.pdf",
                            mime="application/pdf", key="dl_pdf_multi",
                            use_container_width=True)

        # ---- Exportar ----
        st.divider()
        st.subheader("⬇️ Exportar a Excel (lo filtrado)")
        if st.button("Generar Excel", key="exp_dash"):
            out = Path(__file__).parent / "BitaLogs_Export.xlsx"
            cols_exp = ["id"] + COLS_EDIT + ["semana", "dia", "duracion_min"]
            build_excel(fdf[[c for c in cols_exp if c in fdf.columns]],
                        load_comentarios(), out)
            with open(out, "rb") as fh:
                st.download_button(
                    "📥 Descargar Excel", data=fh.read(),
                    file_name="BitaLogs_Export.xlsx",
                    mime="application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet", key="dl_dash")
            st.success("Excel generado.")


# =============================================================== TAB: Datos
with tab_datos:
    df = load_atenciones()
    st.subheader("Todas las atenciones")
    st.caption("Se puede editar cualquier celda, incluidas Semana y Día (por si un "
               "registro se cargó con retraso y quedó en el día equivocado). La "
               "duración se recalcula de las horas. La papelera borra filas.")

    if df.empty:
        st.info("Sin datos todavía.")
    else:
        # Numeración amigable (1,2,3... por Semana→Día), independiente del id.
        df = _orden_secuencial(df)
        # Ocultar las columnas de imagen en el editor (son base64 enormes);
        # se conservan al guardar porque update las preserva desde la BD.
        cols_ocultar = ["img1", "img2", "img3", "img4"]
        show = df.drop(columns=[c for c in cols_ocultar if c in df.columns])
        # Mostrar 'n' (N°) al frente y ordenar la tabla por esa numeración.
        show = show.sort_values("n")
        cols_front = ["n"] + [c for c in show.columns if c != "n"]
        show = show[cols_front]
        show = show.rename(columns={**COLS, "n": "N°"})
        edited = st.data_editor(
            show, use_container_width=True, hide_index=True, num_rows="dynamic",
            key="editor_datos",
            column_config={
                "N°": st.column_config.NumberColumn(
                    "N°", disabled=True,
                    help="Número secuencial de la práctica (no el id interno)."),
                "Área": st.column_config.SelectboxColumn("Área", options=AREAS),
                "Tipo de mantenimiento": st.column_config.SelectboxColumn(
                    "Tipo de mantenimiento", options=TIPOS),
                "¿Resuelto?": st.column_config.SelectboxColumn(
                    "¿Resuelto?", options=RESUELTO),
                "Semana": st.column_config.NumberColumn(
                    "Semana", min_value=1, max_value=10, step=1,
                    help="Editable: corrige la semana del registro."),
                "Día": st.column_config.NumberColumn(
                    "Día", min_value=1, max_value=5, step=1,
                    help="Editable: corrige el día del registro."),
                "Duración (min)": st.column_config.NumberColumn(
                    "Duración (min)", disabled=True),
            },
        )
        if st.button("💾 Guardar cambios", type="primary", key="save_datos"):
            inv = {v: k for k, v in COLS.items()}
            # Quitar la columna de numeración amigable (no existe en la BD).
            edited_save = edited.drop(columns=["N°"], errors="ignore")
            update_atenciones_from_df(edited_save.rename(columns=inv))
            st.success("Cambios guardados.")
            st.rerun()

        st.divider()
        st.subheader("📷 Agregar o cambiar fotos de un registro")
        st.caption("Selecciona un registro (incluidos los antiguos) para agregarle "
                   "hasta 4 fotos. Reemplaza las que tenga.")
        df_ord = df.sort_values("n")
        etqs_img = [f"N° {int(r['n'])}  |  {r['fecha']}  |  S{r['semana']}D{r['dia']}"
                    f"  |  {r['area']}  |  {r['equipo']}"
                    for _, r in df_ord.iterrows()]
        # Mapa etiqueta -> id real de la base
        _map_img = {e: int(r["id"]) for e, (_, r) in zip(etqs_img, df_ord.iterrows())}
        sel_img = st.selectbox("Registro", ["- Selecciona -"] + etqs_img,
                               key="sel_reg_img")
        if sel_img != "- Selecciona -":
            id_img = _map_img[sel_img]
            actuales = imagenes_de(id_img)
            if actuales:
                st.write(f"Este registro ya tiene **{len(actuales)}** foto(s). "
                         "El orden de abajo es el que aparece en la bitácora "
                         "(1 = primera).")
                cols_prev = st.columns(4)
                for i, src in enumerate(actuales):
                    with cols_prev[i]:
                        st.image(src, use_container_width=True)
                        st.caption(f"Posición actual: {i + 1}")

                # ---- Reordenar (#2): asignar nueva posición a cada foto ----
                if len(actuales) > 1:
                    st.markdown("**↕️ Reordenar fotos**")
                    st.caption("Elige la nueva posición de cada foto y guarda. "
                               "Si repites un número, se resuelve por orden.")
                    cols_ord = st.columns(len(actuales))
                    nuevas_pos = []
                    for i, _src in enumerate(actuales):
                        with cols_ord[i]:
                            pos = st.selectbox(
                                f"Foto {i + 1} →",
                                list(range(1, len(actuales) + 1)),
                                index=i, key=f"pos_{id_img}_{i}")
                            nuevas_pos.append(pos)
                    if st.button("💾 Guardar nuevo orden",
                                 key=f"save_orden_{id_img}"):
                        # Ordenar las fotos por la posición pedida; empates
                        # se rompen por el orden original (estable).
                        pares = sorted(
                            zip(nuevas_pos, range(len(actuales)), actuales),
                            key=lambda t: (t[0], t[1]))
                        nuevo = [src for _, _, src in pares]
                        reordenar_imagenes(id_img, nuevo)
                        st.success("✅ Orden de fotos actualizado.")
                        st.rerun()
            else:
                st.write("Este registro **no tiene fotos** todavía.")

            nuevas = st.file_uploader(
                "Nuevas fotos (hasta 4 JPEG), reemplazan las actuales",
                type=["jpg", "jpeg", "png"], accept_multiple_files=True,
                key=f"up_edit_{id_img}")
            colb1, colb2 = st.columns([1, 1])
            with colb1:
                if st.button("💾 Guardar fotos", key=f"save_img_{id_img}",
                             type="primary"):
                    if not nuevas:
                        st.warning("Sube al menos una foto primero.")
                    else:
                        procesadas = []
                        for archivo in nuevas[:4]:
                            b64 = _img_a_base64(archivo)
                            if b64:
                                procesadas.append(b64)
                        if len(nuevas) > 4:
                            st.warning("Solo se guardaron las primeras 4.")
                        update_imagenes(id_img, procesadas)
                        st.success(f"✅ {len(procesadas)} foto(s) guardadas en "
                                   f"el registro #{id_img}.")
                        st.rerun()
            with colb2:
                if actuales and st.button("🗑️ Quitar todas las fotos",
                                          key=f"clear_img_{id_img}"):
                    update_imagenes(id_img, [])
                    st.success(f"Fotos del registro #{id_img} eliminadas.")
                    st.rerun()

        st.divider()
        st.subheader("🗑️ Eliminar una atención")
        df_del = df.sort_values("n")
        etqs = [f"N° {int(r['n'])}  |  {r['fecha']}  |  {r['area']}  |  {r['equipo']}"
                for _, r in df_del.iterrows()]
        _map_del = {e: (int(r["id"]), int(r["n"]))
                    for e, (_, r) in zip(etqs, df_del.iterrows())}
        sel = st.selectbox("Atención a eliminar", ["- Selecciona -"] + etqs,
                           key="del_at")
        if sel != "- Selecciona -":
            id_del, n_del = _map_del[sel]
            if st.checkbox(f"Confirmo eliminar la atención N° {n_del}",
                           key="cf_del"):
                if st.button("Eliminar definitivamente", key="btn_del_at"):
                    delete_atencion(id_del)
                    st.success(f"Atención N° {n_del} eliminada.")
                    st.rerun()


# =============================================================== TAB: Mostrar Bitácoras
with tab_bitacora:
    st.subheader("📄 Mostrar Bitácoras: Matriz de Impacto")
    st.caption("Genera la bitácora con el formato institucional a partir de "
               "los registros. Permite filtrar la tabla y la evidencia "
               "fotográfica por semana o día antes de descargar el HTML.")

    df_b = load_atenciones()
    if df_b.empty:
        st.info("Aún no hay registros para mostrar. Agrega atenciones primero.")
    else:
        # ---- Filtro de la TABLA ----
        st.markdown("**Ver por**")
        cf1, cf2, cf3 = st.columns([1.2, 1, 1])
        with cf1:
            modo = st.radio("Registros", ["Toda la práctica", "Semana", "Día"],
                            horizontal=True, key="bita_modo",
                            label_visibility="collapsed")
        sem_sel = None
        dia_sel = None
        with cf2:
            if modo in ("Semana", "Día"):
                sem_sel = st.selectbox(
                    "Semana", cal.lista_semanas(),
                    format_func=lambda n: f"Semana {n}", key="bita_sem")
        with cf3:
            if modo == "Día":
                dia_sel = st.selectbox(
                    "Día", [1, 2, 3, 4, 5],
                    format_func=lambda d: f"Día {d}", key="bita_dia")

        # ---- Filtro independiente de EVIDENCIA FOTOGRÁFICA ----
        st.markdown("**Evidencia fotográfica**")
        ff1, ff2, ff3 = st.columns([1.2, 1, 1])
        with ff1:
            modo_fotos = st.radio(
                "Fotos", ["Todas", "Por semana", "Por día", "Sin fotos"],
                horizontal=True, key="fotos_modo",
                label_visibility="collapsed",
                help="Elige qué evidencia fotográfica incluir en la bitácora.")
        foto_sem = None
        foto_dia = None
        with ff2:
            if modo_fotos in ("Por semana", "Por día"):
                foto_sem = st.selectbox(
                    "Semana (fotos)", cal.lista_semanas(),
                    format_func=lambda n: f"Semana {n}", key="foto_sem")
        with ff3:
            if modo_fotos == "Por día":
                foto_dia = st.selectbox(
                    "Día (fotos)", [1, 2, 3, 4, 5],
                    format_func=lambda d: f"Día {d}", key="foto_dia")

        # ---- Aplicar filtro de la tabla ----
        dff = df_b.copy()
        titulo = "Matriz de Impacto - Toda la práctica"
        if modo == "Semana" and sem_sel:
            dff = dff[dff["semana"] == sem_sel]
            titulo = f"Matriz de Impacto - Semana {sem_sel}"
        elif modo == "Día" and sem_sel and dia_sel:
            dff = dff[(dff["semana"] == sem_sel) & (dff["dia"] == dia_sel)]
            titulo = f"Matriz de Impacto - Semana {sem_sel}, Día {dia_sel}"

        dff = dff.sort_values(["semana", "dia", "id"], na_position="last")

        # ---- Decidir en qué registros se muestran las fotos ----
        # Copiamos el df y vaciamos las imágenes de los registros que no
        # deban mostrarlas, según el filtro de evidencia (independiente).
        dff_show = dff.copy()
        cols_img = ["img1", "img2", "img3", "img4"]

        def _borrar_fotos(mask):
            for c in cols_img:
                if c in dff_show.columns:
                    dff_show.loc[mask, c] = None

        if modo_fotos == "Sin fotos":
            _borrar_fotos(dff_show.index.notna())  # todas
        elif modo_fotos == "Por semana" and foto_sem:
            _borrar_fotos(dff_show["semana"] != foto_sem)
        elif modo_fotos == "Por día" and foto_sem and foto_dia:
            _borrar_fotos(~((dff_show["semana"] == foto_sem) &
                            (dff_show["dia"] == foto_dia)))
        # "Todas" -> no se borra nada

        # ---- Contadores ----
        n_reg = len(dff_show)
        def _tiene(v):
            return v is not None and str(v).strip() not in ("", "None", "nan")
        n_fotos = 0
        if not dff_show.empty:
            for c in cols_img:
                if c in dff_show.columns:
                    n_fotos += int(dff_show[c].apply(_tiene).sum())
        st.markdown(f"**{n_reg}** registro(s) · **{n_fotos}** imagen(es) "
                    "en la evidencia.")

        registros = dff_show.to_dict("records")
        html = bitacora_html(registros, titulo=titulo)

        st.components.v1.html(html, height=650, scrolling=True)

        # Registros para el Excel: se respeta el filtro de la TABLA (dff),
        # con toda la informacion (el Excel no lleva fotos).
        registros_xlsx = dff.to_dict("records")
        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                "⬇️ Descargar bitácora (HTML para imprimir)",
                data=html.encode("utf-8"),
                file_name=f"{titulo.replace(' ', '_').replace('—','-')}.html",
                mime="text/html", use_container_width=True)
        with dl2:
            st.download_button(
                "⬇️ Descargar Excel (Matriz de impacto)",
                data=matriz_xlsx_bytes(registros_xlsx, titulo=titulo),
                file_name=f"{titulo.replace(' ', '_').replace('—','-')}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet"),
                use_container_width=True)
        st.caption("El HTML incluye la evidencia fotográfica; el Excel replica "
                   "el formato oficial de la Matriz de impacto solo con la "
                   "información de la tabla.")


# =============================================================== TAB: Comentarios
with tab_coment:
    st.subheader("💬 Comentarios semanales de evaluadores")
    st.caption("Los evaluadores pueden dejar un comentario por semana. "
               "Cada comentario queda firmado con su nombre y fecha.")

    with st.form("nuevo_coment", clear_on_submit=True):
        c1, c2 = st.columns([1, 2])
        with c1:
            cs = st.selectbox("Semana", cal.lista_semanas(),
                              format_func=lambda n: cal.etiqueta_semana(n))
        with c2:
            ce = st.text_input("Nombre del evaluador *",
                               placeholder="Ing. Nombre Apellido")
        ct = st.text_area("Comentario *", height=100)
        ok = st.form_submit_button("Guardar comentario", type="primary")
        if ok:
            if not ce.strip() or not ct.strip():
                st.error("El nombre del evaluador y el comentario son obligatorios.")
            else:
                insert_comentario(cs, ce, ct)
                st.success(f"Comentario de {ce.strip()} guardado (Semana {cs}).")

    st.divider()
    dcom = load_comentarios()
    if dcom.empty:
        st.info("Aún no hay comentarios registrados.")
    else:
        for n in cal.lista_semanas():
            sub = dcom[dcom["semana"] == n]
            if sub.empty:
                continue
            st.markdown(f"### {cal.etiqueta_semana(n)}")
            for _, r in sub.iterrows():
                with st.container(border=True):
                    st.markdown(f"**{r['evaluador']}** · _{r['fecha_registro']}_")
                    st.write(r["comentario"])
                    if st.button("🗑️ Eliminar", key=f"delc_{r['id']}"):
                        delete_comentario(int(r["id"]))
                        st.rerun()