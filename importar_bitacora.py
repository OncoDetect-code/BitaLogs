"""
Lector del formato institucional "Matriz de impacto" (UNITEC).

Estructura detectada del archivo:
  - Fila con "Semana N" en la col. A y encabezados en esa misma fila:
        A: Semana N | B: Equipo o proceso | C: Problema identificado
        D: Solución sugerida | E: ¿Problema ya resuelto?
        F: Impacto esperado o beneficio real | G: Observaciones
  - Debajo, filas de datos. La col. A trae "Día 1", "Día 2"... pero solo
    en la PRIMERA fila de cada día (celdas combinadas); las siguientes
    filas del mismo día traen A vacío -> se hereda el día anterior.

Expone:
    leer_matriz(ruta_o_buffer) -> (registros, avisos)
        registros: lista de dicts listos para insertarse en BitaLogs
        avisos:    lista de mensajes (semana no detectada, filas omitidas...)
"""

import re
import openpyxl

# Mapa de "¿resuelto?" del formato -> valores de BitaLogs
_RESUELTO_MAP = {
    "si": "Sí", "sí": "Sí", "yes": "Sí",
    "parcial": "Parcial", "parcialmente": "Parcial",
    "no": "No",
}

_ENC_EQUIPO = "equipo o proceso"


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _num_semana(txt) -> int | None:
    m = re.search(r"semana\s*(\d+)", _norm(txt))
    return int(m.group(1)) if m else None


def _num_dia(txt) -> int | None:
    m = re.search(r"d[ií]a\s*(\d+)", _norm(txt))
    return int(m.group(1)) if m else None


def _map_resuelto(txt) -> str:
    return _RESUELTO_MAP.get(_norm(txt), str(txt or "").strip() or "No")


def leer_matriz(ruta_o_buffer):
    """
    Lee la matriz de impacto y devuelve (registros, avisos).

    Cada registro es un dict con:
        semana, dia, equipo, problema, solucion, resuelto,
        impacto, observaciones
    (fecha, área, horas, etc. las completa el usuario en la app,
     ya que el formato institucional no las trae.)
    """
    wb = openpyxl.load_workbook(ruta_o_buffer, data_only=True)
    ws = wb.active

    registros = []
    avisos = []

    semana_actual = None
    dia_actual = None
    en_datos = False  # estamos debajo de una fila de encabezados

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value

        # ¿Es una fila de encabezado de semana? (A trae "Semana N"
        # y B trae "Equipo o proceso")
        if _num_semana(a) is not None and _norm(b) == _ENC_EQUIPO:
            semana_actual = _num_semana(a)
            dia_actual = None
            en_datos = True
            continue

        if not en_datos:
            continue

        # Actualizar el día si la col. A trae "Día N"
        d = _num_dia(a)
        if d is not None:
            dia_actual = d

        # Fila de datos válida = tiene equipo/proceso en B
        equipo = str(b or "").strip()
        if not equipo:
            continue

        registros.append({
            "semana": semana_actual,
            "dia": dia_actual,
            "equipo": equipo,
            "problema": str(ws.cell(r, 3).value or "").strip(),
            "solucion": str(ws.cell(r, 4).value or "").strip(),
            "resuelto": _map_resuelto(ws.cell(r, 5).value),
            "impacto": str(ws.cell(r, 6).value or "").strip(),
            "observaciones": str(ws.cell(r, 7).value or "").strip(),
        })

    if not registros:
        avisos.append(
            "No se detectaron registros. Verifica que el archivo tenga una "
            "fila con 'Semana N' y encabezados ('Equipo o proceso', etc.).")
    else:
        sin_semana = sum(1 for x in registros if x["semana"] is None)
        if sin_semana:
            avisos.append(
                f"{sin_semana} registro(s) sin número de semana detectado; "
                "podrás asignarlo manualmente al revisar.")

    return registros, avisos


if __name__ == "__main__":
    import sys
    ruta = sys.argv[1] if len(sys.argv) > 1 else \
        "/mnt/user-data/uploads/Matriz_de_impacto_Q3_2026_Sem_1.xlsx"
    regs, avs = leer_matriz(ruta)
    print(f"{len(regs)} registros extraídos:")
    for x in regs:
        print(f"  S{x['semana']} D{x['dia']} · {x['equipo'][:40]:<40} "
              f"resuelto={x['resuelto']}")
    for a in avs:
        print("AVISO:", a)
